import csv
import hashlib
import io
import json
import threading
from datetime import datetime
from uuid import uuid4

from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    File,
    HTTPException,
    Query,
    UploadFile,
)
from openpyxl import load_workbook
from sqlalchemy import desc, func
from sqlalchemy.orm import Session
from sqlalchemy.exc import OperationalError

from ..database import SessionLocal, get_db
from ..models import (
    MeetingReport,
    MeetingReportHighlight,
    MeetingReportQuestion,
    MeetingReportReflection,
    MeetingReportReflectionAudio,
    MeetingReportTranslation,
)
from ..config import settings
from ..schemas import (
    FeishuDocxDiagnoseRequest,
    FeishuDocxDiagnoseResponse,
    FeishuDocxDiagnoseStep,
    FeishuMeetingDiagnoseRequest,
    FeishuMeetingDiagnoseResponse,
    FeishuMeetingDiagnoseStep,
    FeishuDocxInspectRequest,
    FeishuDocxInspectResponse,
    FeishuMeetingImportRequest,
    FeishuMeetingImportItem,
    FeishuMeetingImportResponse,
    FeishuMeetingInspectItem,
    FeishuMeetingInspectRequest,
    FeishuMeetingInspectResponse,
    FeishuDocxImportRequest,
    FeishuDocxImportResponse,
    GeneratePreviewRequest,
    GeneratePreviewResponse,
    GenerateResponse,
    ImportResponse,
    PlaybackModeResponse,
    PlaybackModeUpdateRequest,
    PlaybackQueueItem,
    PlaybackQueueResponse,
    QuestionItem,
    QuestionResponse,
    ReflectionResponse,
    ReflectionItem,
    PublishResponse,
    ReportCreate,
    ReportDetail,
    ReportListItem,
    ReportListResponse,
    ReportTranslationsResponse,
    SynthesizeAudioRequest,
    SynthesizeAudioResponse,
    ReportUpdate,
    SourceImportRequest,
    SourceImportResponse,
    TranslationPrepareResponse,
    TranslationJobStatusItem,
    TranslationJobStatusResponse,
    TranslationJobTriggerResponse,
    TranslationUpdateRequest,
    TranslateScriptRequest,
    TranslateScriptResponse,
)
from ..services.generator import (
    QUESTION_PERSONA_PROMPTS,
    generate_meeting_chapters_from_transcript,
    generate_sharp_questions,
    generate_script_and_highlights,
    normalize_question_persona,
    synthesize_script_audio_pcm_base64,
    translate_report_package,
    translate_script,
)
from ..services.feishu_import import (
    FeishuApiClient,
    FeishuMeetingImportItem as FeishuRawItem,
)
from ..utils.timezone import now_local_naive

router = APIRouter(prefix="/api/reports", tags=["reports"])

LANGUAGE_TARGETS = {
    "zh": "Chinese",
    "en": "English",
    "yue": "Cantonese (Yue Chinese, spoken style in Traditional Chinese characters)",
    "ja": "Japanese",
    "id": "Indonesian",
    "ms": "Malay (Malaysia)",
    "hi": "Hindi",
    "th": "Thai",
}
TEXT_RENDER_LANGUAGE_KEYS = {"zh", "en"}
QUESTION_PERSONA_KEYS = set(QUESTION_PERSONA_PROMPTS.keys())

_translation_job_lock = threading.Lock()
_translation_job_state: dict[tuple[int, str], dict] = {}


def _resolve_render_mode(language_key: str) -> str:
    return "text" if language_key in TEXT_RENDER_LANGUAGE_KEYS else "audio"


def _normalize_source_language(language_key: str | None) -> str:
    key = (language_key or "").strip().lower()
    if key in LANGUAGE_TARGETS:
        return key
    return ""


def _normalize_question_persona_key(persona_key: str | None) -> str:
    key = normalize_question_persona(persona_key)
    if key in QUESTION_PERSONA_KEYS:
        return key
    return "board_director"


def _detect_source_language(title: str, summary_raw: str, script_text: str = "") -> str:
    text = f"{title}\n{summary_raw}\n{script_text}".strip()
    if not text:
        return "zh"

    if any("\u0900" <= ch <= "\u097f" for ch in text):
        return "hi"
    if any("\u3040" <= ch <= "\u30ff" for ch in text):
        return "ja"
    if any("\u4e00" <= ch <= "\u9fff" for ch in text):
        return "zh"

    lower = text.lower()
    ms_hints = ["yang", "dengan", "untuk", "adalah", "dan", "tidak"]
    id_hints = ["yang", "dengan", "untuk", "adalah", "dan", "tidak"]
    ms_score = sum(1 for x in ms_hints if f" {x} " in f" {lower} ")
    id_score = sum(1 for x in id_hints if f" {x} " in f" {lower} ")
    if ms_score >= 3:
        return "ms"
    if id_score >= 3:
        return "id"
    return "en"


def _set_translation_job_state(
    report_id: int, language_key: str, status: str, error: str = ""
) -> None:
    with _translation_job_lock:
        _translation_job_state[(report_id, language_key)] = {
            "status": status,
            "error": error,
            "updated_at": now_local_naive(),
        }


def _get_translation_job_state(report_id: int, language_key: str) -> dict | None:
    with _translation_job_lock:
        state = _translation_job_state.get((report_id, language_key))
        if not state:
            return None
        return {
            "status": str(state.get("status") or ""),
            "error": str(state.get("error") or ""),
            "updated_at": state.get("updated_at"),
        }


def _run_prepare_translation_job(report_id: int, language_key: str) -> None:
    _set_translation_job_state(report_id, language_key, "translating")
    db = SessionLocal()
    try:
        report = db.get(MeetingReport, report_id)
        if report is None:
            _set_translation_job_state(report_id, language_key, "failed", "记录不存在")
            return
        _prepare_translation_row(db, report, language_key)
        db.commit()
        _set_translation_job_state(report_id, language_key, "ready")
    except HTTPException as exc:
        db.rollback()
        _set_translation_job_state(report_id, language_key, "failed", str(exc.detail))
    except Exception as exc:
        db.rollback()
        _set_translation_job_state(
            report_id, language_key, "failed", f"翻译失败: {exc}"
        )
    finally:
        db.close()


def _serialize_report_detail(report: MeetingReport) -> ReportDetail:
    draft = sorted(
        [h for h in report.highlights if h.kind == "draft"], key=lambda x: x.seq
    )
    final = sorted(
        [h for h in report.highlights if h.kind == "final"], key=lambda x: x.seq
    )
    reflections = sorted(report.reflections, key=lambda x: x.seq)
    questions = sorted(report.questions, key=lambda x: x.seq)
    return ReportDetail(
        id=report.id,
        title=report.title,
        meeting_time=report.meeting_time,
        speaker=report.speaker,
        source_language=(report.source_language or "zh"),
        summary_raw=report.summary_raw,
        script_draft=report.script_draft,
        script_final=report.script_final,
        auto_play_enabled=report.auto_play_enabled,
        highlights_draft=[h.highlight_text for h in draft],
        highlights_final=[h.highlight_text for h in final],
        reflections_final=[r.reflection_text for r in reflections],
        questions_final=[q.question_text for q in questions][:3],
        question_persona=_normalize_question_persona_key(report.question_persona),
        status=report.status,
        published_at=report.published_at,
        created_at=report.created_at,
        updated_at=report.updated_at,
    )


def _set_highlights(db: Session, report_id: int, kind: str, values: list[str]) -> None:
    db.query(MeetingReportHighlight).filter(
        MeetingReportHighlight.report_id == report_id,
        MeetingReportHighlight.kind == kind,
    ).delete()
    for idx, text in enumerate(values):
        db.add(
            MeetingReportHighlight(
                report_id=report_id,
                kind=kind,
                highlight_text=text.strip(),
                seq=idx,
            )
        )


def _set_reflections(db: Session, report_id: int, values: list[str]) -> None:
    db.query(MeetingReportReflection).filter(
        MeetingReportReflection.report_id == report_id
    ).delete()
    for idx, text in enumerate(values):
        db.add(
            MeetingReportReflection(
                report_id=report_id,
                reflection_text=text.strip(),
                seq=idx,
            )
        )


def _set_questions(
    db: Session,
    report_id: int,
    values: list[str],
    persona_key: str,
) -> None:
    normalized_persona = _normalize_question_persona_key(persona_key)
    db.query(MeetingReportQuestion).filter(
        MeetingReportQuestion.report_id == report_id
    ).delete()
    for idx, text in enumerate(values[:3]):
        db.add(
            MeetingReportQuestion(
                report_id=report_id,
                question_text=text.strip(),
                persona_key=normalized_persona,
                seq=idx,
            )
        )


def _reflection_text_hash(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _synthesize_reflection_audios_job(
    report_id: int,
    language_keys: list[str] | None = None,
) -> None:
    """后台任务：为每条反思 × 每种语言预合成 TTS 音频并写入 DB。

    language_keys 为 None 时合成所有非 TEXT_RENDER_LANGUAGE_KEYS 的语言；
    传入指定列表时只合成那几种语言。
    zh / en 是文字渲染模式，无需 TTS，跳过。
    """
    from ..services.generator import synthesize_script_audio_pcm_base64

    target_keys = [
        k
        for k in (language_keys or list(LANGUAGE_TARGETS.keys()))
        if k not in TEXT_RENDER_LANGUAGE_KEYS
    ]
    if not target_keys:
        return

    db = SessionLocal()
    try:
        report = db.get(MeetingReport, report_id)
        if report is None:
            return
        reflections = sorted(report.reflections, key=lambda x: x.seq)
        if not reflections:
            return

        for lang_key in target_keys:
            lang_label = LANGUAGE_TARGETS.get(lang_key, lang_key)
            for row in reflections:
                text = str(row.reflection_text).strip()
                if not text:
                    continue
                t_hash = _reflection_text_hash(text)
                # 检查是否已有最新缓存（hash 未变则跳过）
                existing = (
                    db.query(MeetingReportReflectionAudio)
                    .filter(
                        MeetingReportReflectionAudio.report_id == report_id,
                        MeetingReportReflectionAudio.seq == row.seq,
                        MeetingReportReflectionAudio.language_key == lang_key,
                    )
                    .first()
                )
                if existing and existing.text_hash == t_hash:
                    continue  # Stage 4: text_hash 一致，缓存有效，跳过
                try:
                    audio = synthesize_script_audio_pcm_base64(
                        text, lang_key, lang_label
                    )
                except Exception:
                    continue  # 单条失败不影响其他条
                if existing:
                    existing.text_hash = t_hash
                    existing.audio_pcm_base64 = audio
                    existing.updated_at = now_local_naive()
                else:
                    db.add(
                        MeetingReportReflectionAudio(
                            report_id=report_id,
                            seq=row.seq,
                            language_key=lang_key,
                            text_hash=t_hash,
                            audio_pcm_base64=audio,
                        )
                    )
                db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()


def _upsert_translation(
    db: Session,
    report_id: int,
    language_key: str,
    title_text: str,
    script_text: str,
    highlights: list[str],
    reflections: list[str],
    questions: list[str],
    question_persona: str,
    audio_pcm_base64: str = "",
) -> None:
    row = (
        db.query(MeetingReportTranslation)
        .filter(
            MeetingReportTranslation.report_id == report_id,
            MeetingReportTranslation.language_key == language_key,
        )
        .first()
    )
    if row is None:
        row = MeetingReportTranslation(
            report_id=report_id,
            language_key=language_key,
            title_text=title_text.strip(),
            script_text=script_text.strip(),
            highlights_json=json.dumps(highlights[:2], ensure_ascii=False),
            reflections_json=json.dumps(reflections[:5], ensure_ascii=False),
            questions_json=json.dumps(questions[:3], ensure_ascii=False),
            question_persona=_normalize_question_persona_key(question_persona),
            audio_pcm_base64=audio_pcm_base64.strip(),
            updated_at=now_local_naive(),
        )
        db.add(row)
        return

    row.title_text = title_text.strip()
    row.script_text = script_text.strip()
    row.highlights_json = json.dumps(highlights[:2], ensure_ascii=False)
    row.reflections_json = json.dumps(reflections[:5], ensure_ascii=False)
    row.questions_json = json.dumps(questions[:3], ensure_ascii=False)
    row.question_persona = _normalize_question_persona_key(question_persona)
    row.audio_pcm_base64 = audio_pcm_base64.strip()
    row.updated_at = now_local_naive()


def _refresh_report_translations(
    db: Session,
    report: MeetingReport,
    highlights_final: list[str],
) -> None:
    script_text = (report.script_final or "").strip()
    if not script_text:
        return

    title_text = (report.title or "").strip()
    source_highlights = [str(x).strip() for x in highlights_final if str(x).strip()][:2]
    source_reflections = [
        str(x.reflection_text).strip()
        for x in sorted(report.reflections, key=lambda x: x.seq)
        if str(x.reflection_text).strip()
    ][:5]
    source_questions = [
        str(x.question_text).strip()
        for x in sorted(report.questions, key=lambda x: x.seq)
        if str(x.question_text).strip()
    ][:3]
    source_persona = _normalize_question_persona_key(report.question_persona)

    source_lang = _normalize_source_language(
        report.source_language
    ) or _detect_source_language(report.title, report.summary_raw, script_text)
    translated_payloads: dict[
        str, tuple[str, str, list[str], list[str], list[str], str, str]
    ] = {
        source_lang: (
            title_text,
            script_text,
            source_highlights,
            source_reflections,
            source_questions,
            source_persona,
            "",
        ),
    }

    for language_key, target_language in LANGUAGE_TARGETS.items():
        if language_key == source_lang:
            continue
        try:
            title_out, script_out, highlights_out = translate_report_package(
                title_text=title_text,
                script_text=script_text,
                highlights=source_highlights,
                target_language=target_language,
            )
            reflections_out: list[str] = []
            for reflection in source_reflections:
                try:
                    reflections_out.append(
                        translate_script(reflection, target_language)
                    )
                except Exception:
                    reflections_out.append(reflection)
            questions_out: list[str] = []
            for question in source_questions:
                try:
                    questions_out.append(translate_script(question, target_language))
                except Exception:
                    questions_out.append(question)
            audio_pcm_base64 = ""
            if _resolve_render_mode(language_key) == "audio":
                try:
                    audio_pcm_base64 = synthesize_script_audio_pcm_base64(
                        script_text=script_out,
                        language_key=language_key,
                        language_label=target_language,
                    )
                except Exception:
                    # 音频失败时先写入翻译文本，后续异步任务可重试。
                    audio_pcm_base64 = ""
            translated_payloads[language_key] = (
                title_out,
                script_out,
                highlights_out,
                reflections_out[:5],
                questions_out[:3],
                source_persona,
                audio_pcm_base64,
            )
        except Exception:
            # 单语种失败不影响保存；保留已有翻译或后续重试。
            continue

    for language_key, payload in translated_payloads.items():
        (
            title_out,
            script_out,
            highlights_out,
            reflections_out,
            questions_out,
            question_persona,
            audio_pcm_base64,
        ) = payload
        _upsert_translation(
            db=db,
            report_id=report.id,
            language_key=language_key,
            title_text=title_out,
            script_text=script_out,
            highlights=highlights_out,
            reflections=reflections_out,
            questions=questions_out,
            question_persona=question_persona,
            audio_pcm_base64=audio_pcm_base64,
        )


def _refresh_report_translations_job(report_id: int) -> None:
    db = SessionLocal()
    try:
        report = db.get(MeetingReport, report_id)
        if report is None:
            return
        if not (report.script_final or "").strip():
            return
        final_rows = sorted(
            [h for h in report.highlights if h.kind == "final"], key=lambda x: x.seq
        )
        final_highlights = [h.highlight_text for h in final_rows][:2]
        _refresh_report_translations(db, report, final_highlights)
        db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()


def _normalize_highlights(raw: list[str] | None) -> list[str]:
    if not raw:
        return []
    return [str(x).strip() for x in raw if str(x).strip()][:2]


def _normalize_reflections(raw: list[str] | None) -> list[str]:
    if not raw:
        return []
    cleaned = [str(x).strip() for x in raw if str(x).strip()]
    return cleaned[:5]


def _normalize_questions(raw: list[str] | None) -> list[str]:
    if not raw:
        return []
    cleaned = [str(x).strip() for x in raw if str(x).strip()]
    return cleaned[:3]


def _parse_translation_highlights(row: MeetingReportTranslation) -> list[str]:
    try:
        highlights = json.loads(row.highlights_json) if row.highlights_json else []
        if not isinstance(highlights, list):
            return []
        return [str(x).strip() for x in highlights if str(x).strip()][:2]
    except Exception:
        return []


def _parse_translation_reflections(row: MeetingReportTranslation) -> list[str]:
    try:
        reflections = json.loads(row.reflections_json) if row.reflections_json else []
        if not isinstance(reflections, list):
            return []
        return [str(x).strip() for x in reflections if str(x).strip()][:5]
    except Exception:
        return []


def _parse_translation_questions(row: MeetingReportTranslation) -> list[str]:
    try:
        questions = json.loads(row.questions_json) if row.questions_json else []
        if not isinstance(questions, list):
            return []
        return [str(x).strip() for x in questions if str(x).strip()][:3]
    except Exception:
        return []


def _build_feishu_summary(
    source: FeishuRawItem, ai_chapters: list[str] | None = None
) -> str:
    lines: list[str] = [
        f"会议号: {source.meeting_no}",
        f"会议实例ID: {source.meeting_id}",
        f"妙记链接: {source.minute_url or '未获取'}",
        f"文字记录状态: {source.transcript_status}",
    ]
    if source.transcript_error:
        lines.append(f"文字记录说明: {source.transcript_error}")
    if source.minute_summary_formal:
        lines.append("")
        lines.append("【会议章节总结（正式）】")
        lines.append(source.minute_summary_formal)
    if ai_chapters:
        cleaned = [str(x).strip() for x in ai_chapters if str(x).strip()][:6]
        if cleaned:
            lines.append("")
            lines.append("【AI章节纪要】")
            lines.extend(cleaned)
    if source.transcript_text:
        lines.append("")
        lines.append("【会议文字记录】")
        lines.append(source.transcript_text)
    if source.minute_summary_draft and not source.minute_summary_formal:
        lines.append("")
        lines.append("【会议实时草稿】")
        lines.append(source.minute_summary_draft)
    return "\n".join(lines).strip()


def _ensure_report_script(report: MeetingReport, db: Session) -> tuple[str, list[str]]:
    script = (report.script_final or report.script_draft or "").strip()
    final_highlights_rows = sorted(
        [h for h in report.highlights if h.kind == "final"], key=lambda x: x.seq
    )
    final_highlights = [h.highlight_text for h in final_highlights_rows][:2]

    if script:
        return script, _normalize_highlights(final_highlights)

    (
        generated_script,
        generated_highlights,
        generated_reflections,
        generated_questions,
    ) = generate_script_and_highlights(
        summary_raw=report.summary_raw,
        speaker=report.speaker,
        title=report.title,
        question_persona=_normalize_question_persona_key(report.question_persona),
    )
    report.script_draft = generated_script
    report.script_final = generated_script
    _set_highlights(db, report.id, "draft", generated_highlights)
    _set_highlights(db, report.id, "final", generated_highlights)
    _set_reflections(db, report.id, _normalize_reflections(generated_reflections))
    _set_questions(
        db,
        report.id,
        _normalize_questions(generated_questions),
        _normalize_question_persona_key(report.question_persona),
    )
    report.updated_at = now_local_naive()
    db.flush()
    return generated_script, generated_highlights[:2]


def _prepare_translation_row(
    db: Session, report: MeetingReport, language_key: str
) -> MeetingReportTranslation:
    if language_key not in LANGUAGE_TARGETS:
        raise HTTPException(status_code=400, detail="不支持的语言")

    base_script, base_highlights = _ensure_report_script(report, db)
    base_reflections = [
        str(x.reflection_text).strip()
        for x in sorted(report.reflections, key=lambda x: x.seq)
        if str(x.reflection_text).strip()
    ][:5]
    base_questions = [
        str(x.question_text).strip()
        for x in sorted(report.questions, key=lambda x: x.seq)
        if str(x.question_text).strip()
    ][:3]
    base_persona = _normalize_question_persona_key(report.question_persona)
    base_title = (report.title or "").strip()
    target_language = LANGUAGE_TARGETS[language_key]
    source_lang = _normalize_source_language(
        report.source_language
    ) or _detect_source_language(report.title, report.summary_raw, base_script)

    if language_key == source_lang:
        title_out = base_title
        script_out = base_script
        highlights_out = base_highlights
        reflections_out = base_reflections
        questions_out = base_questions
        question_persona = base_persona
        audio_pcm_base64 = ""
        if _resolve_render_mode(language_key) == "audio" and script_out.strip():
            try:
                audio_pcm_base64 = synthesize_script_audio_pcm_base64(
                    script_text=script_out,
                    language_key=language_key,
                    language_label=target_language,
                )
            except Exception:
                audio_pcm_base64 = ""
    else:
        title_out, script_out, highlights_out = translate_report_package(
            title_text=base_title,
            script_text=base_script,
            highlights=base_highlights,
            target_language=target_language,
        )
        reflections_out: list[str] = []
        for reflection in base_reflections:
            try:
                reflections_out.append(translate_script(reflection, target_language))
            except Exception:
                reflections_out.append(reflection)
        questions_out: list[str] = []
        for question in base_questions:
            try:
                questions_out.append(translate_script(question, target_language))
            except Exception:
                questions_out.append(question)
        question_persona = base_persona
        audio_pcm_base64 = ""
        if _resolve_render_mode(language_key) == "audio":
            audio_pcm_base64 = synthesize_script_audio_pcm_base64(
                script_text=script_out,
                language_key=language_key,
                language_label=target_language,
            )

    row = (
        db.query(MeetingReportTranslation)
        .filter(
            MeetingReportTranslation.report_id == report.id,
            MeetingReportTranslation.language_key == language_key,
        )
        .first()
    )
    if row is None:
        row = MeetingReportTranslation(
            report_id=report.id,
            language_key=language_key,
            title_text=title_out,
            script_text=script_out,
            highlights_json=json.dumps(highlights_out[:2], ensure_ascii=False),
            reflections_json=json.dumps(reflections_out[:5], ensure_ascii=False),
            questions_json=json.dumps(questions_out[:3], ensure_ascii=False),
            question_persona=question_persona,
            audio_pcm_base64=audio_pcm_base64,
            reviewed=False,
            reviewed_at=None,
            updated_at=now_local_naive(),
        )
        db.add(row)
        db.flush()
        return row

    row.title_text = title_out
    row.script_text = script_out
    row.highlights_json = json.dumps(highlights_out[:2], ensure_ascii=False)
    row.reflections_json = json.dumps(reflections_out[:5], ensure_ascii=False)
    row.questions_json = json.dumps(questions_out[:3], ensure_ascii=False)
    row.question_persona = question_persona
    row.audio_pcm_base64 = audio_pcm_base64
    row.reviewed = False
    row.reviewed_at = None
    row.updated_at = now_local_naive()
    db.flush()
    return row


def _translation_response_item(
    row: MeetingReportTranslation, report: MeetingReport
) -> TranslationPrepareResponse:
    highlights = _parse_translation_highlights(row)
    reflections = _parse_translation_reflections(row)
    questions = _parse_translation_questions(row)
    language_key = row.language_key
    render_mode = _resolve_render_mode(language_key)
    return TranslationPrepareResponse(
        report_id=report.id,
        language_key=language_key,
        status="ready",
        reviewed=bool(row.reviewed),
        reviewed_at=row.reviewed_at,
        title=(row.title_text or report.title or "").strip(),
        script_final=(row.script_text or report.script_final or "").strip(),
        highlights_final=highlights,
        reflections_final=reflections,
        questions_final=questions,
        question_persona=_normalize_question_persona_key(row.question_persona),
        render_mode=render_mode,
        audio_ready=True
        if render_mode == "text"
        else bool((row.audio_pcm_base64 or "").strip()),
    )


def _build_localized_payload(
    db: Session,
    report: MeetingReport,
    highlights_final: list[str],
    reflections_final: list[str],
    questions_final: list[str],
    include_audio: bool = False,
    audio_languages: set[str] | None = None,
) -> dict[str, dict]:
    audio_langs = audio_languages or set()
    include_all_audio = not audio_langs
    source_lang = _normalize_source_language(
        report.source_language
    ) or _detect_source_language(report.title, report.summary_raw, report.script_final)
    payload: dict[str, dict] = {}
    rows = (
        db.query(MeetingReportTranslation)
        .filter(MeetingReportTranslation.report_id == report.id)
        .all()
    )
    row_by_lang = {row.language_key: row for row in rows}
    source_render_mode = _resolve_render_mode(source_lang)
    source_row = row_by_lang.get(source_lang)
    source_audio_ready = (
        True
        if source_render_mode == "text"
        else bool((source_row.audio_pcm_base64 or "").strip())
        if source_row
        else False
    )
    source_audio_pcm = ""
    if (
        source_row
        and source_render_mode == "audio"
        and include_audio
        and (include_all_audio or source_lang in audio_langs)
    ):
        source_audio_pcm = (source_row.audio_pcm_base64 or "").strip()
    payload[source_lang] = {
        "title": report.title,
        "script_final": report.script_final,
        "highlights_final": highlights_final[:2],
        "reflections_final": reflections_final[:5],
        "questions_final": questions_final[:3],
        "question_persona": _normalize_question_persona_key(report.question_persona),
        "render_mode": source_render_mode,
        "audio_ready": source_audio_ready,
        "audio_pcm_base64": source_audio_pcm,
    }
    for row in rows:
        try:
            highlights = json.loads(row.highlights_json) if row.highlights_json else []
            if not isinstance(highlights, list):
                highlights = []
        except Exception:
            highlights = []
        reflections = _parse_translation_reflections(row)
        questions = _parse_translation_questions(row)
        render_mode = _resolve_render_mode(row.language_key)
        audio_pcm_base64 = ""
        audio_ready = True
        if render_mode == "audio":
            if include_audio and (include_all_audio or row.language_key in audio_langs):
                audio_pcm_base64 = (row.audio_pcm_base64 or "").strip()
            audio_ready = bool((row.audio_pcm_base64 or "").strip())
        payload[row.language_key] = {
            "title": row.title_text or report.title,
            "script_final": row.script_text or report.script_final,
            "highlights_final": [str(x).strip() for x in highlights if str(x).strip()][
                :2
            ],
            "reflections_final": reflections if reflections else reflections_final[:5],
            "questions_final": questions if questions else questions_final[:3],
            "question_persona": _normalize_question_persona_key(row.question_persona),
            "render_mode": render_mode,
            "audio_ready": audio_ready,
            "audio_pcm_base64": audio_pcm_base64,
        }
    return payload


@router.get("", response_model=ReportListResponse)
def list_reports(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status: str | None = Query(None),
    db: Session = Depends(get_db),
):
    query = db.query(MeetingReport)
    if status:
        query = query.filter(MeetingReport.status == status)

    total = query.with_entities(func.count(MeetingReport.id)).scalar() or 0
    items = (
        query.order_by(desc(MeetingReport.meeting_time), desc(MeetingReport.id))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return ReportListResponse(
        items=[
            ReportListItem(
                id=item.id,
                title=item.title,
                speaker=item.speaker,
                source_language=(item.source_language or "zh"),
                meeting_time=item.meeting_time,
                status=item.status,
                auto_play_enabled=item.auto_play_enabled,
            )
            for item in items
        ],
        total=total,
        page=page,
        page_size=page_size,
    )


@router.post("", response_model=ReportDetail)
def create_report(
    payload: ReportCreate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
):
    if not payload.title.strip():
        raise HTTPException(status_code=400, detail="标题不能为空")
    if not payload.summary_raw.strip():
        raise HTTPException(status_code=400, detail="内容不能为空")

    meeting_time = payload.meeting_time or now_local_naive()
    speaker = payload.speaker.strip() if payload.speaker else ""
    source_language = _normalize_source_language(
        payload.source_language
    ) or _detect_source_language(
        payload.title, payload.summary_raw, payload.script_final
    )
    report = MeetingReport(
        title=payload.title,
        meeting_time=meeting_time,
        speaker=speaker,
        summary_raw=payload.summary_raw,
        source_language=source_language,
        question_persona=_normalize_question_persona_key(payload.question_persona),
        script_final=payload.script_final,
        auto_play_enabled=payload.auto_play_enabled,
        status="draft",
        created_at=now_local_naive(),
        updated_at=now_local_naive(),
    )
    db.add(report)
    db.flush()

    if payload.highlights_final:
        _set_highlights(db, report.id, "final", payload.highlights_final[:2])
    if payload.reflections_final:
        _set_reflections(
            db, report.id, _normalize_reflections(payload.reflections_final)
        )
    if payload.questions_final:
        _set_questions(
            db,
            report.id,
            _normalize_questions(payload.questions_final),
            _normalize_question_persona_key(payload.question_persona),
        )

    db.commit()
    db.refresh(report)
    if (report.script_final or "").strip():
        background_tasks.add_task(_refresh_report_translations_job, report.id)
    if payload.reflections_final:
        background_tasks.add_task(_synthesize_reflection_audios_job, report.id)
    return _serialize_report_detail(report)


@router.get("/published/latest", response_model=ReportDetail)
def get_latest_published(db: Session = Depends(get_db)):
    report = (
        db.query(MeetingReport)
        .filter(MeetingReport.status == "published")
        .order_by(desc(MeetingReport.published_at), desc(MeetingReport.id))
        .first()
    )
    if not report:
        raise HTTPException(status_code=404, detail="暂无已发布内容")
    return _serialize_report_detail(report)


@router.get("/{report_id}", response_model=ReportDetail)
def get_report(report_id: int, db: Session = Depends(get_db)):
    report = db.get(MeetingReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="记录不存在")
    return _serialize_report_detail(report)


@router.get("/{report_id}/translations", response_model=ReportTranslationsResponse)
def get_report_translations(report_id: int, db: Session = Depends(get_db)):
    report = db.get(MeetingReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="记录不存在")

    rows = (
        db.query(MeetingReportTranslation)
        .filter(MeetingReportTranslation.report_id == report.id)
        .all()
    )
    row_by_lang = {row.language_key: row for row in rows}

    base_highlights_rows = sorted(
        [h for h in report.highlights if h.kind == "final"], key=lambda x: x.seq
    )
    base_highlights = [h.highlight_text for h in base_highlights_rows][:2]
    base_reflections_rows = sorted(report.reflections, key=lambda x: x.seq)
    base_reflections = [x.reflection_text for x in base_reflections_rows][:5]
    base_questions_rows = sorted(report.questions, key=lambda x: x.seq)
    base_questions = [x.question_text for x in base_questions_rows][:3]
    base_question_persona = _normalize_question_persona_key(report.question_persona)
    source_lang = _normalize_source_language(
        report.source_language
    ) or _detect_source_language(report.title, report.summary_raw, report.script_final)
    items: list = []
    for language_key in LANGUAGE_TARGETS.keys():
        render_mode = _resolve_render_mode(language_key)
        job_state = _get_translation_job_state(report.id, language_key)
        status_override = str(job_state.get("status") or "") if job_state else ""
        job_error = str(job_state.get("error") or "") if job_state else ""

        if language_key == source_lang:
            script_text = (report.script_final or report.script_draft or "").strip()
            items.append(
                {
                    "language_key": language_key,
                    "status": "ready" if script_text else "missing",
                    "error": "",
                    "reviewed": True,
                    "reviewed_at": report.updated_at,
                    "title": report.title,
                    "script_final": script_text,
                    "highlights_final": base_highlights,
                    "reflections_final": base_reflections,
                    "questions_final": base_questions,
                    "question_persona": base_question_persona,
                    "render_mode": render_mode,
                    "audio_ready": True,
                }
            )
            continue

        row = row_by_lang.get(language_key)
        if not row:
            items.append(
                {
                    "language_key": language_key,
                    "status": status_override if status_override else "missing",
                    "error": job_error,
                    "reviewed": False,
                    "reviewed_at": None,
                    "title": "",
                    "script_final": "",
                    "highlights_final": [],
                    "reflections_final": [],
                    "questions_final": [],
                    "question_persona": base_question_persona,
                    "render_mode": render_mode,
                    "audio_ready": False if render_mode == "audio" else True,
                }
            )
            continue

        highlights = _parse_translation_highlights(row)
        reflections = _parse_translation_reflections(row)
        questions = _parse_translation_questions(row)
        status = "ready"
        if status_override in {"translating", "failed"}:
            status = status_override
        items.append(
            {
                "language_key": language_key,
                "status": status,
                "error": job_error,
                "reviewed": bool(row.reviewed),
                "reviewed_at": row.reviewed_at,
                "title": row.title_text or report.title,
                "script_final": row.script_text or report.script_final,
                "highlights_final": highlights,
                "reflections_final": reflections if reflections else base_reflections,
                "questions_final": questions if questions else base_questions,
                "question_persona": _normalize_question_persona_key(
                    row.question_persona
                ),
                "render_mode": render_mode,
                "audio_ready": True
                if render_mode == "text"
                else bool((row.audio_pcm_base64 or "").strip()),
            }
        )

    return ReportTranslationsResponse(report_id=report.id, items=items)


@router.get(
    "/{report_id}/translation-jobs", response_model=TranslationJobStatusResponse
)
def get_report_translation_jobs(report_id: int, db: Session = Depends(get_db)):
    report = db.get(MeetingReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="记录不存在")

    source_lang = _normalize_source_language(
        report.source_language
    ) or _detect_source_language(report.title, report.summary_raw, report.script_final)
    items: list[TranslationJobStatusItem] = []
    for language_key in LANGUAGE_TARGETS.keys():
        if language_key == source_lang:
            continue
        state = _get_translation_job_state(report.id, language_key)
        if not state:
            items.append(
                TranslationJobStatusItem(
                    language_key=language_key,
                    status="missing",
                    error="",
                    updated_at=None,
                )
            )
            continue
        items.append(
            TranslationJobStatusItem(
                language_key=language_key,
                status=str(state.get("status") or "missing"),
                error=str(state.get("error") or ""),
                updated_at=state.get("updated_at"),
            )
        )
    return TranslationJobStatusResponse(report_id=report.id, items=items)


@router.post(
    "/{report_id}/translations/retranslate-all",
    response_model=TranslationJobTriggerResponse,
)
def retranslate_all_languages(
    report_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
):
    report = db.get(MeetingReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="记录不存在")

    source_lang = _normalize_source_language(
        report.source_language
    ) or _detect_source_language(report.title, report.summary_raw, report.script_final)
    target_languages = [x for x in LANGUAGE_TARGETS.keys() if x != source_lang]
    if not target_languages:
        raise HTTPException(status_code=400, detail="没有可翻译的目标语种")

    task_id = uuid4().hex
    for language_key in target_languages:
        _set_translation_job_state(report.id, language_key, "translating")
        background_tasks.add_task(_run_prepare_translation_job, report.id, language_key)

    return TranslationJobTriggerResponse(
        task_id=task_id,
        report_id=report.id,
        languages=target_languages,
        status="queued",
    )


@router.post(
    "/{report_id}/translations/{language_key}/retranslate",
    response_model=TranslationJobTriggerResponse,
)
def retranslate_single_language(
    report_id: int,
    language_key: str,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
):
    if language_key not in LANGUAGE_TARGETS:
        raise HTTPException(status_code=400, detail="不支持的语言")

    report = db.get(MeetingReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="记录不存在")

    source_lang = _normalize_source_language(
        report.source_language
    ) or _detect_source_language(report.title, report.summary_raw, report.script_final)
    if language_key == source_lang:
        raise HTTPException(status_code=400, detail="原始语种主稿无需重译")

    task_id = uuid4().hex
    _set_translation_job_state(report.id, language_key, "translating")
    background_tasks.add_task(_run_prepare_translation_job, report.id, language_key)

    return TranslationJobTriggerResponse(
        task_id=task_id,
        report_id=report.id,
        languages=[language_key],
        status="queued",
    )


@router.post(
    "/{report_id}/translations/{language_key}/prepare",
    response_model=TranslationPrepareResponse,
)
def prepare_report_translation(
    report_id: int, language_key: str, db: Session = Depends(get_db)
):
    report = db.get(MeetingReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="记录不存在")

    try:
        row = _prepare_translation_row(db, report, language_key)
        db.commit()
        db.refresh(report)
        db.refresh(row)
    except HTTPException:
        raise
    except ValueError as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"翻译准备失败: {exc}") from exc

    return _translation_response_item(row, report)


@router.put(
    "/{report_id}/translations/{language_key}",
    response_model=TranslationPrepareResponse,
)
def update_report_translation(
    report_id: int,
    language_key: str,
    payload: TranslationUpdateRequest,
    db: Session = Depends(get_db),
):
    if language_key not in LANGUAGE_TARGETS:
        raise HTTPException(status_code=400, detail="不支持的语言")

    report = db.get(MeetingReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="记录不存在")
    source_lang = _normalize_source_language(
        report.source_language
    ) or _detect_source_language(report.title, report.summary_raw, report.script_final)
    if language_key == source_lang:
        raise HTTPException(
            status_code=400, detail="原始语种主稿请在编辑页主稿区域保存"
        )

    row = (
        db.query(MeetingReportTranslation)
        .filter(
            MeetingReportTranslation.report_id == report.id,
            MeetingReportTranslation.language_key == language_key,
        )
        .first()
    )
    if row is None:
        try:
            row = _prepare_translation_row(db, report, language_key)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    data = payload.model_dump(exclude_unset=True)
    if "title" in data and data["title"] is not None:
        row.title_text = data["title"].strip()
    if "script_final" in data and data["script_final"] is not None:
        row.script_text = data["script_final"].strip()
        if _resolve_render_mode(language_key) == "audio":
            try:
                row.audio_pcm_base64 = synthesize_script_audio_pcm_base64(
                    script_text=row.script_text,
                    language_key=language_key,
                    language_label=LANGUAGE_TARGETS[language_key],
                )
            except Exception:
                row.audio_pcm_base64 = ""
    if "highlights_final" in data and data["highlights_final"] is not None:
        row.highlights_json = json.dumps(
            _normalize_highlights(data["highlights_final"]), ensure_ascii=False
        )
    if "reflections_final" in data and data["reflections_final"] is not None:
        row.reflections_json = json.dumps(
            _normalize_reflections(data["reflections_final"]), ensure_ascii=False
        )
    if "questions_final" in data and data["questions_final"] is not None:
        row.questions_json = json.dumps(
            _normalize_questions(data["questions_final"]), ensure_ascii=False
        )
    if "question_persona" in data and data["question_persona"] is not None:
        row.question_persona = _normalize_question_persona_key(data["question_persona"])
    if "reviewed" in data and data["reviewed"] is not None:
        row.reviewed = bool(data["reviewed"])
        row.reviewed_at = now_local_naive() if row.reviewed else None

    row.updated_at = now_local_naive()
    db.commit()
    db.refresh(report)
    db.refresh(row)
    return _translation_response_item(row, report)


@router.put("/{report_id}", response_model=ReportDetail)
def update_report(
    report_id: int,
    payload: ReportUpdate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
):
    report = db.get(MeetingReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="记录不存在")

    data = payload.model_dump(exclude_unset=True)
    previous_title = (report.title or "").strip()
    previous_script = (report.script_final or "").strip()
    previous_source_language = (report.source_language or "").strip().lower()
    previous_highlights = [
        h.highlight_text
        for h in sorted(
            [x for x in report.highlights if x.kind == "final"], key=lambda x: x.seq
        )
    ]
    previous_reflections = [
        x.reflection_text for x in sorted(report.reflections, key=lambda x: x.seq)
    ]
    previous_questions = [
        x.question_text for x in sorted(report.questions, key=lambda x: x.seq)
    ]

    for key in [
        "title",
        "meeting_time",
        "speaker",
        "summary_raw",
        "script_final",
        "auto_play_enabled",
    ]:
        if key in data:
            setattr(report, key, data[key])
    if "source_language" in data:
        normalized = _normalize_source_language(data["source_language"])
        if not normalized:
            raise HTTPException(status_code=400, detail="不支持的原始语种")
        report.source_language = normalized
    elif not _normalize_source_language(report.source_language):
        report.source_language = _detect_source_language(
            report.title, report.summary_raw, report.script_final
        )
    if "question_persona" in data and data["question_persona"] is not None:
        report.question_persona = _normalize_question_persona_key(
            data["question_persona"]
        )

    highlights_changed = False
    if "highlights_final" in data and data["highlights_final"] is not None:
        highlights = [x.strip() for x in data["highlights_final"] if x.strip()]
        if len(highlights) > 2:
            raise HTTPException(status_code=400, detail="highlights_final 最多 2 条")
        if highlights != previous_highlights:
            _set_highlights(db, report.id, "final", highlights)
            highlights_changed = True
    reflections_changed = False
    if "reflections_final" in data and data["reflections_final"] is not None:
        raw_reflections = [
            str(x).strip() for x in data["reflections_final"] if str(x).strip()
        ]
        if len(raw_reflections) > 5:
            raise HTTPException(status_code=400, detail="reflections_final 最多 5 条")
        normalized_reflections = _normalize_reflections(raw_reflections)
        if normalized_reflections != previous_reflections:
            _set_reflections(db, report.id, normalized_reflections)
            reflections_changed = True
    questions_changed = False
    if "questions_final" in data and data["questions_final"] is not None:
        raw_questions = [
            str(x).strip() for x in data["questions_final"] if str(x).strip()
        ]
        if len(raw_questions) > 3:
            raise HTTPException(status_code=400, detail="questions_final 最多 3 条")
        normalized_questions = _normalize_questions(raw_questions)
        if normalized_questions != previous_questions:
            _set_questions(
                db,
                report.id,
                normalized_questions,
                _normalize_question_persona_key(report.question_persona),
            )
            questions_changed = True

    if not report.title.strip():
        raise HTTPException(status_code=400, detail="标题不能为空")
    if not report.summary_raw.strip():
        raise HTTPException(status_code=400, detail="内容不能为空")

    title_changed = previous_title != (report.title or "").strip()
    script_changed = previous_script != (report.script_final or "").strip()
    source_language_changed = (
        previous_source_language != (report.source_language or "").strip().lower()
    )
    should_refresh_translation = any(
        [
            title_changed,
            script_changed,
            highlights_changed,
            reflections_changed,
            questions_changed,
            source_language_changed,
        ]
    )
    report.updated_at = now_local_naive()
    try:
        db.commit()
    except OperationalError as exc:
        db.rollback()
        if "database is locked" in str(exc).lower():
            raise HTTPException(
                status_code=503, detail="数据库忙，请稍后重试（建议 1-2 秒后再次保存）"
            ) from exc
        raise
    db.refresh(report)
    if should_refresh_translation and (report.script_final or "").strip():
        background_tasks.add_task(_refresh_report_translations_job, report.id)
    if reflections_changed:
        background_tasks.add_task(_synthesize_reflection_audios_job, report.id)
    return _serialize_report_detail(report)


@router.delete("/{report_id}")
def delete_report(report_id: int, db: Session = Depends(get_db)):
    report = db.get(MeetingReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="记录不存在")
    db.delete(report)
    db.commit()
    return {"ok": True, "deleted_id": report_id}


@router.post("/{report_id}/generate", response_model=GenerateResponse)
def generate_report_content(
    report_id: int, background_tasks: BackgroundTasks, db: Session = Depends(get_db)
):
    report = db.get(MeetingReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="记录不存在")
    if not report.summary_raw.strip():
        raise HTTPException(status_code=400, detail="内容不能为空")
    if not _normalize_source_language(report.source_language):
        report.source_language = _detect_source_language(
            report.title, report.summary_raw, report.script_final
        )

    try:
        script, highlights, reflections, questions = generate_script_and_highlights(
            report.summary_raw,
            report.speaker,
            report.title,
            question_persona=_normalize_question_persona_key(report.question_persona),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    report.script_draft = script
    report.script_final = script
    report.updated_at = now_local_naive()
    _set_highlights(db, report.id, "draft", highlights[:2])
    _set_highlights(db, report.id, "final", highlights[:2])
    _set_reflections(db, report.id, _normalize_reflections(reflections))
    _set_questions(
        db,
        report.id,
        _normalize_questions(questions),
        _normalize_question_persona_key(report.question_persona),
    )
    db.commit()
    background_tasks.add_task(_refresh_report_translations_job, report.id)
    background_tasks.add_task(_synthesize_reflection_audios_job, report.id)

    return GenerateResponse(
        report_id=report.id,
        script_draft=script,
        highlights_draft=highlights[:2],
        reflections_draft=_normalize_reflections(reflections),
        questions_draft=_normalize_questions(questions),
    )


@router.post("/{report_id}/generate-pack", response_model=GenerateResponse)
def generate_report_pack(report_id: int, db: Session = Depends(get_db)):
    report = db.get(MeetingReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="记录不存在")
    if not report.summary_raw.strip():
        raise HTTPException(status_code=400, detail="内容不能为空")
    if not _normalize_source_language(report.source_language):
        report.source_language = _detect_source_language(
            report.title, report.summary_raw, report.script_final
        )

    try:
        script, highlights, reflections, questions = generate_script_and_highlights(
            report.summary_raw,
            report.speaker,
            report.title,
            question_persona=_normalize_question_persona_key(report.question_persona),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    report.script_draft = script
    report.script_final = script
    report.updated_at = now_local_naive()
    _set_highlights(db, report.id, "draft", highlights[:2])
    _set_highlights(db, report.id, "final", highlights[:2])
    _set_reflections(db, report.id, _normalize_reflections(reflections))
    _set_questions(
        db,
        report.id,
        _normalize_questions(questions),
        _normalize_question_persona_key(report.question_persona),
    )
    db.flush()

    # 统一生成入口：同步准备多语言（包含反思），避免前端看到“已生成但未准备”。
    _refresh_report_translations(db, report, highlights[:2])
    db.commit()

    return GenerateResponse(
        report_id=report.id,
        script_draft=script,
        highlights_draft=highlights[:2],
        reflections_draft=_normalize_reflections(reflections),
        questions_draft=_normalize_questions(questions),
    )


@router.post("/generate-preview", response_model=GeneratePreviewResponse)
def generate_preview_content(payload: GeneratePreviewRequest):
    if not payload.title.strip():
        raise HTTPException(status_code=400, detail="标题不能为空")
    if not payload.summary_raw.strip():
        raise HTTPException(status_code=400, detail="内容不能为空")

    try:
        script, highlights, reflections, questions = generate_script_and_highlights(
            payload.summary_raw,
            payload.speaker,
            payload.title,
            question_persona=_normalize_question_persona_key(payload.question_persona),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return GeneratePreviewResponse(
        script_draft=script,
        highlights_draft=highlights[:2],
        reflections_draft=_normalize_reflections(reflections),
        questions_draft=_normalize_questions(questions),
    )


@router.post("/translate-script", response_model=TranslateScriptResponse)
def translate_report_script(payload: TranslateScriptRequest):
    try:
        translated = translate_script(payload.script_text, payload.target_language)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return TranslateScriptResponse(translated_text=translated)


@router.post("/synthesize-audio", response_model=SynthesizeAudioResponse)
def synthesize_report_audio(payload: SynthesizeAudioRequest):
    script_text = (payload.script_text or "").strip()
    language_key = (payload.language_key or "").strip().lower()
    if not script_text:
        raise HTTPException(status_code=400, detail="script_text 不能为空")
    if language_key not in LANGUAGE_TARGETS:
        raise HTTPException(status_code=400, detail="不支持的 language_key")

    try:
        audio_pcm_base64 = synthesize_script_audio_pcm_base64(
            script_text=script_text,
            language_key=language_key,
            language_label=LANGUAGE_TARGETS[language_key],
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"AI 语音合成失败: {exc}") from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"语音合成失败: {exc}") from exc

    return SynthesizeAudioResponse(
        language_key=language_key,
        audio_pcm_base64=audio_pcm_base64,
    )


@router.get("/{report_id}/reflection", response_model=ReflectionResponse)
def get_report_reflection(
    report_id: int,
    lang: str = Query(default="zh"),
    db: Session = Depends(get_db),
):
    report = db.get(MeetingReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="记录不存在")
    reflections = sorted(report.reflections, key=lambda x: x.seq)

    # 预加载该语言所有已合成音频，以 seq 为 key
    audio_rows = (
        db.query(MeetingReportReflectionAudio)
        .filter(
            MeetingReportReflectionAudio.report_id == report_id,
            MeetingReportReflectionAudio.language_key == lang,
        )
        .all()
    )
    audio_by_seq: dict[int, MeetingReportReflectionAudio] = {
        r.seq: r for r in audio_rows
    }

    items: list[ReflectionItem] = []
    for row in reflections:
        text = str(row.reflection_text).strip()
        t_hash = _reflection_text_hash(text)
        audio_row = audio_by_seq.get(row.seq)
        # Stage 4: text_hash 不一致说明文字已变动，缓存失效
        audio_pcm = (
            audio_row.audio_pcm_base64
            if audio_row and audio_row.text_hash == t_hash
            else None
        )
        items.append(ReflectionItem(text=text, audio_pcm_base64=audio_pcm))

    return ReflectionResponse(report_id=report.id, reflections=items)


@router.get("/{report_id}/questions", response_model=QuestionResponse)
def get_report_questions(
    report_id: int,
    lang: str = Query(default="zh"),
    persona: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    report = db.get(MeetingReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="记录不存在")

    target_persona = _normalize_question_persona_key(persona or report.question_persona)
    source_persona = _normalize_question_persona_key(report.question_persona)
    report_questions = sorted(report.questions, key=lambda x: x.seq)
    base_questions = [
        str(x.question_text).strip()
        for x in report_questions
        if str(x.question_text).strip()
    ][:3]

    if not base_questions:
        base_questions = generate_sharp_questions(
            title_text=report.title,
            summary_text=report.summary_raw,
            script_text=report.script_final,
            persona_key=target_persona,
        )
        _set_questions(
            db, report.id, _normalize_questions(base_questions), target_persona
        )
        report.question_persona = target_persona
        report.updated_at = now_local_naive()
        db.commit()
        db.refresh(report)
        source_persona = target_persona

    if target_persona != source_persona:
        generated = generate_sharp_questions(
            title_text=report.title,
            summary_text=report.summary_raw,
            script_text=report.script_final,
            persona_key=target_persona,
        )
        return QuestionResponse(
            report_id=report.id,
            persona=target_persona,
            questions=[QuestionItem(text=x) for x in _normalize_questions(generated)],
        )

    if lang.strip().lower() != _normalize_source_language(report.source_language):
        row = (
            db.query(MeetingReportTranslation)
            .filter(
                MeetingReportTranslation.report_id == report.id,
                MeetingReportTranslation.language_key == lang.strip().lower(),
            )
            .first()
        )
        if row:
            translated_questions = _parse_translation_questions(row)
            if translated_questions:
                return QuestionResponse(
                    report_id=report.id,
                    persona=_normalize_question_persona_key(row.question_persona),
                    questions=[QuestionItem(text=x) for x in translated_questions],
                )

    return QuestionResponse(
        report_id=report.id,
        persona=source_persona,
        questions=[QuestionItem(text=x) for x in _normalize_questions(base_questions)],
    )


@router.post("/{report_id}/publish", response_model=PublishResponse)
def publish_report(report_id: int, db: Session = Depends(get_db)):
    report = db.get(MeetingReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="记录不存在")

    final_highlights = sorted(
        [h.highlight_text for h in report.highlights if h.kind == "final"]
    )
    if not report.title.strip() or not report.speaker.strip():
        raise HTTPException(status_code=400, detail="标题和发言人不能为空")
    if not report.script_final.strip():
        raise HTTPException(status_code=400, detail="script_final 不能为空")
    if not (1 <= len(final_highlights) <= 2):
        raise HTTPException(status_code=400, detail="highlights_final 必须是 1 到 2 条")

    report.status = "published"
    report.published_at = now_local_naive()
    report.updated_at = now_local_naive()
    db.commit()

    return PublishResponse(
        report_id=report.id,
        status=report.status,
        published_at=report.published_at or now_local_naive(),
    )


@router.post("/import/file", response_model=ImportResponse)
async def import_reports_from_file(
    file: UploadFile = File(...), db: Session = Depends(get_db)
):
    filename = (file.filename or "").lower()
    content = await file.read()
    rows: list[dict] = []

    if filename.endswith(".csv"):
        decoded = content.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(decoded))
        rows = [dict(r) for r in reader]
    elif filename.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="XLSX 工作表为空")
        headers = [
            str(c.value).strip() if c.value else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        for row in ws.iter_rows(min_row=2):
            item = {}
            for idx, cell in enumerate(row):
                key = headers[idx] if idx < len(headers) else f"col_{idx}"
                item[key] = "" if cell.value is None else str(cell.value)
            rows.append(item)
    else:
        raise HTTPException(status_code=400, detail="仅支持 CSV 或 XLSX 文件")

    success_count = 0
    failed_count = 0
    errors: list[str] = []

    for idx, row in enumerate(rows, start=2):
        try:
            title = (row.get("title") or row.get("标题") or "").strip()
            speaker = (row.get("speaker") or row.get("发言人") or "").strip()
            summary_raw = (row.get("summary_raw") or row.get("总结原文") or "").strip()
            meeting_time_str = (
                row.get("meeting_time") or row.get("时间") or ""
            ).strip()

            if not title or not speaker or not summary_raw or not meeting_time_str:
                raise ValueError("缺少必填字段")

            try:
                meeting_time = datetime.fromisoformat(meeting_time_str)
            except ValueError:
                meeting_time = datetime.strptime(meeting_time_str, "%Y-%m-%d")

            db.add(
                MeetingReport(
                    title=title,
                    speaker=speaker,
                    summary_raw=summary_raw,
                    source_language=_detect_source_language(title, summary_raw, ""),
                    meeting_time=meeting_time,
                    auto_play_enabled=False,
                    status="draft",
                    created_at=now_local_naive(),
                    updated_at=now_local_naive(),
                )
            )
            success_count += 1
        except Exception as exc:
            failed_count += 1
            errors.append(f"第 {idx} 行导入失败: {exc}")

    db.commit()
    return ImportResponse(
        success_count=success_count, failed_count=failed_count, errors=errors
    )


@router.post("/import/source", response_model=SourceImportResponse)
def import_reports_from_source(payload: SourceImportRequest):
    # 预留第三方 API 导入能力，后续接异步任务队列。
    return SourceImportResponse(
        task_id=f"pending-{uuid4().hex}",
        accepted_count=0,
        message="已接收导入请求，当前为预留接口，后续版本接入实际拉取逻辑。",
    )


@router.post(
    "/import/feishu-meeting/inspect", response_model=FeishuMeetingInspectResponse
)
def inspect_feishu_meeting(payload: FeishuMeetingInspectRequest):
    if not settings.feishu_app_id or not settings.feishu_app_secret:
        raise HTTPException(
            status_code=400, detail="后端未配置 FEISHU_APP_ID / FEISHU_APP_SECRET"
        )

    try:
        client = FeishuApiClient(
            app_id=settings.feishu_app_id,
            app_secret=settings.feishu_app_secret,
            api_base=settings.feishu_api_base,
            timeout_sec=settings.feishu_timeout_sec,
            verify_ssl=settings.feishu_verify_ssl,
        )
        source_items = client.fetch_items_from_url(
            source_url=payload.meeting_url,
            lookback_days=payload.lookback_days,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(
            status_code=500, detail=f"飞书会议信息查询失败: {exc}"
        ) from exc

    items: list[FeishuMeetingInspectItem] = []
    for source in source_items:
        items.append(
            FeishuMeetingInspectItem(
                meeting_no=source.meeting_no,
                meeting_id=source.meeting_id,
                topic=source.topic,
                start_time=source.start_time,
                end_time=source.end_time,
                minute_token=source.minute_token,
                minute_url=source.minute_url,
                minute_title=source.minute_title,
                minute_owner_id=source.minute_owner_id,
                minute_summary_formal=source.minute_summary_formal,
                minute_summary_draft=source.minute_summary_draft,
                transcript_status=source.transcript_status,
                transcript_error=source.transcript_error,
                transcript_text=source.transcript_text,
            )
        )

    return FeishuMeetingInspectResponse(
        total=len(items),
        items=items,
        message="飞书会议信息查询完成",
    )


@router.post(
    "/import/feishu-meeting/diagnose", response_model=FeishuMeetingDiagnoseResponse
)
def diagnose_feishu_meeting(payload: FeishuMeetingDiagnoseRequest):
    if not settings.feishu_app_id or not settings.feishu_app_secret:
        raise HTTPException(
            status_code=400, detail="后端未配置 FEISHU_APP_ID / FEISHU_APP_SECRET"
        )

    steps: list[FeishuMeetingDiagnoseStep] = []
    meeting_no = ""
    suggestion = "可读取 transcript，可继续导入并自动生成。"

    try:
        client = FeishuApiClient(
            app_id=settings.feishu_app_id,
            app_secret=settings.feishu_app_secret,
            api_base=settings.feishu_api_base,
            timeout_sec=settings.feishu_timeout_sec,
            verify_ssl=settings.feishu_verify_ssl,
        )
    except Exception as exc:
        raise HTTPException(
            status_code=500, detail=f"飞书客户端初始化失败: {exc}"
        ) from exc

    minute_token_from_url = client.parse_minutes_token(payload.meeting_url)
    if minute_token_from_url:
        meeting_no = ""
        steps.append(
            FeishuMeetingDiagnoseStep(
                step="parse_minutes_token",
                ok=True,
                detail="妙记链接解析成功",
                data={"minute_token": minute_token_from_url},
            )
        )
        minute_payload = client.get_minute(minute_token=minute_token_from_url)
        minute_ok = minute_payload.get("code") == 0
        steps.append(
            FeishuMeetingDiagnoseStep(
                step="minute_get",
                ok=minute_ok,
                detail="妙记详情可读"
                if minute_ok
                else f"妙记详情不可读: {minute_payload.get('msg') or 'unknown'}",
                data={
                    "code": minute_payload.get("code"),
                    "msg": minute_payload.get("msg"),
                },
            )
        )
        transcript_status, transcript_text, transcript_error = client.get_transcript(
            minute_token=minute_token_from_url
        )
        transcript_ok = transcript_status == "ready" and bool(transcript_text.strip())
        steps.append(
            FeishuMeetingDiagnoseStep(
                step="minute_transcript",
                ok=transcript_ok,
                detail="妙记转写可读"
                if transcript_ok
                else f"妙记转写不可读: {transcript_status}",
                data={
                    "transcript_status": transcript_status,
                    "transcript_error": transcript_error,
                    "transcript_length": len(transcript_text or ""),
                },
            )
        )
        overall_ok = all(x.ok for x in steps)
        if not overall_ok:
            suggestion = "妙记链接已识别，但读取权限不足。请检查 minutes 相关权限并完成企业安装授权。"
        return FeishuMeetingDiagnoseResponse(
            ok=overall_ok,
            meeting_no="",
            steps=steps,
            suggestion=suggestion,
        )
    try:
        meeting_no = client.parse_meeting_no(payload.meeting_url)
        steps.append(
            FeishuMeetingDiagnoseStep(
                step="parse_meeting_no",
                ok=True,
                detail="会议链接解析成功",
                data={"meeting_no": meeting_no},
            )
        )
    except Exception as exc:
        steps.append(
            FeishuMeetingDiagnoseStep(
                step="parse_meeting_no",
                ok=False,
                detail=f"会议链接解析失败: {exc}",
                data={},
            )
        )
        return FeishuMeetingDiagnoseResponse(
            ok=False,
            meeting_no="",
            steps=steps,
            suggestion="请检查链接格式，支持会议链接和妙记链接。",
        )

    from datetime import datetime, timedelta

    end_time = int(datetime.now().timestamp())
    start_time = int(
        (
            datetime.now() - timedelta(days=max(1, int(payload.lookback_days)))
        ).timestamp()
    )

    try:
        briefs = client.list_meetings_by_no(
            meeting_no=meeting_no, start_ts=start_time, end_ts=end_time
        )
        steps.append(
            FeishuMeetingDiagnoseStep(
                step="list_by_no",
                ok=True,
                detail="会议列表查询成功",
                data={"count": len(briefs)},
            )
        )
    except Exception as exc:
        steps.append(
            FeishuMeetingDiagnoseStep(
                step="list_by_no",
                ok=False,
                detail=f"会议列表查询失败: {exc}",
                data={"lookback_days": payload.lookback_days},
            )
        )
        return FeishuMeetingDiagnoseResponse(
            ok=False,
            meeting_no=meeting_no,
            steps=steps,
            suggestion="请确认会议号、时间窗口和应用会议查询权限是否已发布并完成企业安装授权。",
        )

    if not briefs:
        steps.append(
            FeishuMeetingDiagnoseStep(
                step="pick_meeting",
                ok=False,
                detail="未查询到任何会议实例",
                data={},
            )
        )
        return FeishuMeetingDiagnoseResponse(
            ok=False,
            meeting_no=meeting_no,
            steps=steps,
            suggestion="请增大 lookback_days 或确认该会议属于当前租户。",
        )

    meeting_id = str(briefs[0].get("id") or "").strip()
    if not meeting_id:
        steps.append(
            FeishuMeetingDiagnoseStep(
                step="pick_meeting",
                ok=False,
                detail="会议实例缺少 meeting_id",
                data={"brief": briefs[0]},
            )
        )
        return FeishuMeetingDiagnoseResponse(
            ok=False,
            meeting_no=meeting_no,
            steps=steps,
            suggestion="请确认会议数据完整性，或更换会议链接重试。",
        )

    steps.append(
        FeishuMeetingDiagnoseStep(
            step="pick_meeting",
            ok=True,
            detail="已选取会议实例",
            data={"meeting_id": meeting_id},
        )
    )

    detail = client.get_meeting_detail(meeting_id=meeting_id)
    steps.append(
        FeishuMeetingDiagnoseStep(
            step="meeting_detail",
            ok=bool(detail),
            detail="会议详情查询完成" if detail else "会议详情为空（可能权限不足）",
            data={
                "topic": str(detail.get("topic") or ""),
                "status": detail.get("status"),
                "start_time": detail.get("start_time"),
                "end_time": detail.get("end_time"),
            },
        )
    )

    recording_payload = client.get_recording(meeting_id=meeting_id)
    recording = recording_payload.get("recording") or {}
    recording_url = str(recording.get("url") or "").strip()
    steps.append(
        FeishuMeetingDiagnoseStep(
            step="meeting_recording",
            ok=bool(recording_url),
            detail="录制信息查询完成"
            if recording_url
            else f"录制信息不可用: {recording_payload.get('msg') or 'unknown'}",
            data={
                "recording_url": recording_url,
                "code": recording_payload.get("code"),
            },
        )
    )

    minute_token = client.parse_minutes_token(recording_url) if recording_url else ""
    steps.append(
        FeishuMeetingDiagnoseStep(
            step="parse_minute_token",
            ok=bool(minute_token),
            detail="minute_token 解析成功" if minute_token else "minute_token 解析失败",
            data={"minute_token": minute_token},
        )
    )
    if not minute_token:
        meeting_status = str(detail.get("status") or "").strip()
        recording_code = recording_payload.get("code")
        recording_msg = str(recording_payload.get("msg") or "").strip().lower()
        if (
            meeting_status == "2"
            or recording_code == 122001
            or recording_msg == "meeting status unexpected"
        ):
            return FeishuMeetingDiagnoseResponse(
                ok=True,
                meeting_no=meeting_no,
                steps=steps,
                suggestion="会议进行中，录制/妙记暂未产出。已兼容该场景，可先绑定并展示占位实时记录，待会议结束后再自动获取正式纪要。",
            )
        return FeishuMeetingDiagnoseResponse(
            ok=False,
            meeting_no=meeting_no,
            steps=steps,
            suggestion="会议录制或妙记尚未生成，请稍后重试；或确认会议已开启录制与妙记。",
        )

    minute_payload = client.get_minute(minute_token=minute_token)
    minute_ok = minute_payload.get("code") == 0
    steps.append(
        FeishuMeetingDiagnoseStep(
            step="minute_get",
            ok=minute_ok,
            detail="妙记详情可读"
            if minute_ok
            else f"妙记详情不可读: {minute_payload.get('msg') or 'unknown'}",
            data={"code": minute_payload.get("code"), "msg": minute_payload.get("msg")},
        )
    )

    transcript_status, transcript_text, transcript_error = client.get_transcript(
        minute_token=minute_token
    )
    transcript_ok = transcript_status == "ready" and bool(transcript_text.strip())
    steps.append(
        FeishuMeetingDiagnoseStep(
            step="minute_transcript",
            ok=transcript_ok,
            detail="妙记转写可读"
            if transcript_ok
            else f"妙记转写不可读: {transcript_status}",
            data={
                "transcript_status": transcript_status,
                "transcript_error": transcript_error,
                "transcript_length": len(transcript_text or ""),
            },
        )
    )

    overall_ok = all(
        x.ok for x in steps[-3:]
    )  # minute_token/minute_get/minute_transcript
    if not overall_ok:
        suggestion = "会议可读但妙记权限不足。请在飞书应用补充妙记读取/转写读取权限，发布新版本并重新安装授权后重试。"

    return FeishuMeetingDiagnoseResponse(
        ok=overall_ok,
        meeting_no=meeting_no,
        steps=steps,
        suggestion=suggestion,
    )


@router.post("/import/feishu-docx/inspect", response_model=FeishuDocxInspectResponse)
def inspect_feishu_docx(payload: FeishuDocxInspectRequest):
    if not settings.feishu_app_id or not settings.feishu_app_secret:
        raise HTTPException(
            status_code=400, detail="后端未配置 FEISHU_APP_ID / FEISHU_APP_SECRET"
        )

    client = FeishuApiClient(
        app_id=settings.feishu_app_id,
        app_secret=settings.feishu_app_secret,
        api_base=settings.feishu_api_base,
        timeout_sec=settings.feishu_timeout_sec,
        verify_ssl=settings.feishu_verify_ssl,
    )
    document_id = client.parse_docx_token(payload.docx_url)
    if not document_id:
        return FeishuDocxInspectResponse(
            ok=False,
            document_id="",
            title="",
            content_length=0,
            message="文档链接解析失败",
            error="请使用 https://*.feishu.cn/docx/<token> 格式链接",
        )

    result = client.get_docx_raw_content(document_id)
    if result.get("code") != 0:
        return FeishuDocxInspectResponse(
            ok=False,
            document_id=document_id,
            title="",
            content_length=0,
            message="读取文档内容失败",
            error=f"飞书API错误: {result.get('msg') or 'unknown'} (code: {result.get('code')})",
        )

    text_content = FeishuApiClient.extract_docx_text(result.get("content") or {})
    lines = [line.strip() for line in text_content.split("\n") if line.strip()]
    title = lines[0][:100] if lines else f"飞书文档_{document_id[:8]}"
    content_length = len(text_content.strip())
    if content_length <= 0:
        return FeishuDocxInspectResponse(
            ok=False,
            document_id=document_id,
            title=title,
            content_length=0,
            message="文档内容为空",
            error="读取成功但未提取到文本内容",
        )

    return FeishuDocxInspectResponse(
        ok=True,
        document_id=document_id,
        title=title,
        content_length=content_length,
        message="文档检查通过，可导入",
        error=None,
    )


@router.post("/import/feishu-docx/diagnose", response_model=FeishuDocxDiagnoseResponse)
def diagnose_feishu_docx(payload: FeishuDocxDiagnoseRequest):
    if not settings.feishu_app_id or not settings.feishu_app_secret:
        raise HTTPException(
            status_code=400, detail="后端未配置 FEISHU_APP_ID / FEISHU_APP_SECRET"
        )

    client = FeishuApiClient(
        app_id=settings.feishu_app_id,
        app_secret=settings.feishu_app_secret,
        api_base=settings.feishu_api_base,
        timeout_sec=settings.feishu_timeout_sec,
        verify_ssl=settings.feishu_verify_ssl,
    )

    steps: list[FeishuDocxDiagnoseStep] = []
    suggestion = "文档可读，可继续导入并自动生成。"
    document_id = client.parse_docx_token(payload.docx_url)
    steps.append(
        FeishuDocxDiagnoseStep(
            step="parse_docx_token",
            ok=bool(document_id),
            detail="文档链接解析成功" if document_id else "文档链接解析失败",
            data={"document_id": document_id},
        )
    )
    if not document_id:
        return FeishuDocxDiagnoseResponse(
            ok=False,
            document_id="",
            steps=steps,
            suggestion="请检查链接格式，支持 https://*.feishu.cn/docx/<token>。",
        )

    raw_payload = client.get_docx_raw_content(document_id)
    raw_ok = raw_payload.get("code") == 0
    steps.append(
        FeishuDocxDiagnoseStep(
            step="docx_raw_content",
            ok=raw_ok,
            detail="文档原始内容可读"
            if raw_ok
            else f"文档原始内容不可读: {raw_payload.get('msg') or 'unknown'}",
            data={"code": raw_payload.get("code"), "msg": raw_payload.get("msg")},
        )
    )

    extracted_text = (
        FeishuApiClient.extract_docx_text(raw_payload.get("content") or {})
        if raw_ok
        else ""
    )
    text_ok = bool(extracted_text.strip())
    steps.append(
        FeishuDocxDiagnoseStep(
            step="docx_text_extract",
            ok=text_ok,
            detail="文档文本提取成功" if text_ok else "文档文本提取为空",
            data={"content_length": len(extracted_text or "")},
        )
    )

    overall_ok = all(step.ok for step in steps)
    if not overall_ok:
        suggestion = "请检查文档访问权限，确认飞书应用已授权 docx/drive 只读权限并完成企业安装授权。"

    return FeishuDocxDiagnoseResponse(
        ok=overall_ok,
        document_id=document_id,
        steps=steps,
        suggestion=suggestion,
    )


@router.post("/import/feishu-docx", response_model=FeishuDocxImportResponse)
def import_report_from_feishu_docx(
    payload: FeishuDocxImportRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
):
    """从飞书文档链接导入内容并自动生成口播稿、Highlights和反思问答

    支持的链接格式：
    - https://tpc.feishu.cn/docx/xxxxxx
    """
    if not settings.feishu_app_id or not settings.feishu_app_secret:
        raise HTTPException(
            status_code=400, detail="后端未配置 FEISHU_APP_ID / FEISHU_APP_SECRET"
        )

    # 1. 解析文档token
    try:
        client = FeishuApiClient(
            app_id=settings.feishu_app_id,
            app_secret=settings.feishu_app_secret,
            api_base=settings.feishu_api_base,
            timeout_sec=settings.feishu_timeout_sec,
            verify_ssl=settings.feishu_verify_ssl,
        )
        document_id = client.parse_docx_token(payload.docx_url)
        if not document_id:
            raise ValueError(f"无法从URL解析文档token: {payload.docx_url}")
    except ValueError as exc:
        return FeishuDocxImportResponse(
            success=False,
            report_id=None,
            title="",
            content_length=0,
            message="文档链接解析失败",
            error=str(exc),
        )

    # 2. 读取文档内容
    try:
        result = client.get_docx_raw_content(document_id)
        if result["code"] != 0:
            error_msg = result.get("msg") or "未知错误"
            return FeishuDocxImportResponse(
                success=False,
                report_id=None,
                title="",
                content_length=0,
                message="读取文档内容失败",
                error=f"飞书API错误: {error_msg} (code: {result['code']})",
            )

        # 提取纯文本
        raw_content = result.get("content") or {}
        text_content = FeishuApiClient.extract_docx_text(raw_content)

        if not text_content.strip():
            return FeishuDocxImportResponse(
                success=False,
                report_id=None,
                title="",
                content_length=0,
                message="文档内容为空",
                error="提取的文本内容为空，请检查文档是否有内容",
            )
    except Exception as exc:
        return FeishuDocxImportResponse(
            success=False,
            report_id=None,
            title="",
            content_length=0,
            message="读取文档失败",
            error=str(exc),
        )

    # 3. 创建新闻记录
    title = f"飞书文档_{document_id[:8]}"  # 默认标题
    try:
        # 尝试从文档内容提取标题（取第一行作为标题）
        lines = [line.strip() for line in text_content.split("\n") if line.strip()]
        if lines:
            title = lines[0][:100]  # 限制标题长度
    except Exception:
        pass

    report = MeetingReport(
        title=title,
        meeting_time=now_local_naive(),
        speaker="",
        summary_raw=text_content,
        source_language=_detect_source_language(title, text_content, ""),
        question_persona="board_director",
        script_draft="",
        script_final="",
        auto_play_enabled=bool(payload.auto_enable_playback),
        status="draft",
        source_type="feishu_docx",
        source_meeting_no="",
        source_meeting_id=f"docx:{document_id}",
        source_minute_token="",
        source_url=payload.docx_url,
        created_at=now_local_naive(),
        updated_at=now_local_naive(),
    )
    db.add(report)
    db.commit()
    db.refresh(report)

    # 4. 自动生成（如果启用）
    if payload.auto_generate and text_content.strip():
        try:
            script, highlights, reflections, questions = generate_script_and_highlights(
                summary_raw=report.summary_raw,
                speaker=report.speaker,
                title=report.title,
                question_persona=_normalize_question_persona_key(
                    report.question_persona
                ),
            )
            report.script_draft = script
            report.script_final = script
            report.updated_at = now_local_naive()
            _set_highlights(db, report.id, "draft", highlights[:2])
            _set_highlights(db, report.id, "final", highlights[:2])
            _set_reflections(db, report.id, _normalize_reflections(reflections))
            _set_questions(
                db,
                report.id,
                _normalize_questions(questions),
                _normalize_question_persona_key(report.question_persona),
            )
            db.commit()
            background_tasks.add_task(_refresh_report_translations_job, report.id)
            background_tasks.add_task(_synthesize_reflection_audios_job, report.id)
        except Exception:
            # 生成失败不阻断导入链路，保留原始文档内容供人工编辑
            db.rollback()

    return FeishuDocxImportResponse(
        success=True,
        report_id=report.id,
        title=title,
        content_length=len(text_content),
        message=f"文档导入成功，已创建新闻 ID: {report.id}",
        error=None,
    )


@router.post("/import/feishu-meeting", response_model=FeishuMeetingImportResponse)
def import_reports_from_feishu_meeting(
    payload: FeishuMeetingImportRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
):
    if not settings.feishu_app_id or not settings.feishu_app_secret:
        raise HTTPException(
            status_code=400, detail="后端未配置 FEISHU_APP_ID / FEISHU_APP_SECRET"
        )

    try:
        client = FeishuApiClient(
            app_id=settings.feishu_app_id,
            app_secret=settings.feishu_app_secret,
            api_base=settings.feishu_api_base,
            timeout_sec=settings.feishu_timeout_sec,
            verify_ssl=settings.feishu_verify_ssl,
        )
        source_items = client.fetch_items_from_url(
            source_url=payload.meeting_url,
            lookback_days=payload.lookback_days,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"飞书导入失败: {exc}") from exc

    # 同一次导入按 meeting_id 去重，避免重复处理。
    unique_items: list[FeishuRawItem] = []
    seen_meeting_ids: set[str] = set()
    for item in source_items:
        if item.meeting_id in seen_meeting_ids:
            continue
        seen_meeting_ids.add(item.meeting_id)
        unique_items.append(item)

    imported_count = 0
    updated_count = 0
    failed_count = 0
    response_items: list[FeishuMeetingImportItem] = []
    translation_refresh_ids: set[int] = set()

    for source in unique_items:
        try:
            title = (
                source.minute_title or source.topic or f"飞书会议 {source.meeting_no}"
            ).strip()
            ai_chapters: list[str] = []
            if (
                source.transcript_status == "ready"
                and source.transcript_text.strip()
                and not source.minute_summary_formal.strip()
            ):
                try:
                    ai_chapters = generate_meeting_chapters_from_transcript(
                        title_text=title, transcript_text=source.transcript_text
                    )
                except Exception:
                    ai_chapters = []
            summary_raw = _build_feishu_summary(source, ai_chapters=ai_chapters)
            speaker = (source.minute_owner_id or "").strip()
            meeting_time = source.start_time or now_local_naive()

            report = None
            if source.minute_token:
                report = (
                    db.query(MeetingReport)
                    .filter(MeetingReport.source_minute_token == source.minute_token)
                    .first()
                )
            if report is None:
                report = (
                    db.query(MeetingReport)
                    .filter(
                        MeetingReport.source_type == "feishu_meeting",
                        MeetingReport.source_meeting_id == source.meeting_id,
                    )
                    .first()
                )

            created = report is None
            if created:
                report = MeetingReport(
                    title=title,
                    meeting_time=meeting_time,
                    speaker=speaker,
                    summary_raw=summary_raw,
                    source_language=_detect_source_language(title, summary_raw, ""),
                    question_persona="board_director",
                    script_draft="",
                    script_final="",
                    auto_play_enabled=bool(payload.auto_enable_playback),
                    status="draft",
                    source_type="feishu_meeting",
                    source_meeting_no=source.meeting_no,
                    source_meeting_id=source.meeting_id,
                    source_minute_token=source.minute_token,
                    source_url=source.minute_url or payload.meeting_url,
                    created_at=now_local_naive(),
                    updated_at=now_local_naive(),
                )
                db.add(report)
                db.flush()
            if report is None:
                raise RuntimeError("报告创建失败")
            if not created:
                report.title = title
                report.meeting_time = meeting_time
                report.speaker = speaker
                report.summary_raw = summary_raw
                if not _normalize_source_language(report.source_language):
                    report.source_language = _detect_source_language(
                        title, summary_raw, report.script_final
                    )
                report.source_type = "feishu_meeting"
                report.source_meeting_no = source.meeting_no
                report.source_meeting_id = source.meeting_id
                report.source_minute_token = source.minute_token
                report.source_url = source.minute_url or payload.meeting_url
                if payload.auto_enable_playback:
                    report.auto_play_enabled = True
                report.updated_at = now_local_naive()
                db.flush()

            note_parts: list[str] = []
            should_generate = bool(
                payload.auto_generate
                and source.transcript_status == "ready"
                and source.transcript_text.strip()
            )
            if should_generate:
                try:
                    script, highlights, reflections, questions = (
                        generate_script_and_highlights(
                            summary_raw=report.summary_raw,
                            speaker=report.speaker,
                            title=report.title,
                            question_persona=_normalize_question_persona_key(
                                report.question_persona
                            ),
                        )
                    )
                    report.script_draft = script
                    report.script_final = script
                    _set_highlights(db, report.id, "draft", highlights[:2])
                    _set_highlights(db, report.id, "final", highlights[:2])
                    _set_reflections(db, report.id, _normalize_reflections(reflections))
                    _set_questions(
                        db,
                        report.id,
                        _normalize_questions(questions),
                        _normalize_question_persona_key(report.question_persona),
                    )
                    translation_refresh_ids.add(report.id)
                    note_parts.append("已自动生成口播稿与亮点")
                except Exception as exc:
                    note_parts.append(f"AI 生成失败: {exc}")
            else:
                if payload.auto_generate:
                    note_parts.append(
                        f"未执行 AI 生成（文字记录状态: {source.transcript_status}）"
                    )

            if source.transcript_error:
                note_parts.append(source.transcript_error)
            if ai_chapters:
                note_parts.append(f"已基于逐字稿生成章节纪要 {len(ai_chapters)} 条")

            if created:
                imported_count += 1
            else:
                updated_count += 1

            response_items.append(
                FeishuMeetingImportItem(
                    meeting_no=source.meeting_no,
                    meeting_id=source.meeting_id,
                    minute_token=source.minute_token,
                    minute_url=source.minute_url,
                    title=report.title,
                    report_id=report.id,
                    transcript_status=source.transcript_status,
                    note="；".join([x for x in note_parts if x]),
                )
            )
        except Exception as exc:
            failed_count += 1
            response_items.append(
                FeishuMeetingImportItem(
                    meeting_no=source.meeting_no,
                    meeting_id=source.meeting_id,
                    minute_token=source.minute_token,
                    minute_url=source.minute_url,
                    title=source.minute_title or source.topic,
                    report_id=None,
                    transcript_status=source.transcript_status,
                    note=f"导入失败: {exc}",
                )
            )

    db.commit()
    for report_id in translation_refresh_ids:
        background_tasks.add_task(_refresh_report_translations_job, report_id)
        background_tasks.add_task(_synthesize_reflection_audios_job, report_id)

    message = f"飞书会议导入完成：新增 {imported_count}，更新 {updated_count}，失败 {failed_count}"
    return FeishuMeetingImportResponse(
        imported_count=imported_count,
        updated_count=updated_count,
        failed_count=failed_count,
        items=response_items,
        message=message,
    )


@router.get("/playback/queue", response_model=PlaybackQueueResponse)
def get_playback_queue(
    include_audio: bool = Query(False),
    langs: str | None = Query(None),
    report_id: int | None = Query(None),
    db: Session = Depends(get_db),
):
    langs_text = langs if isinstance(langs, str) else ""
    audio_languages = {x.strip() for x in langs_text.split(",") if x.strip()}
    report_id_value = report_id if isinstance(report_id, int) else None
    reports = (
        db.query(MeetingReport)
        .filter(MeetingReport.auto_play_enabled.is_(True))
        .order_by(desc(MeetingReport.meeting_time), desc(MeetingReport.id))
        .all()
    )
    if report_id_value is not None:
        reports = [x for x in reports if x.id == report_id_value]

    items: list[PlaybackQueueItem] = []
    for report in reports:
        if not report.script_final.strip():
            continue
        highlights_final = sorted(
            [h for h in report.highlights if h.kind == "final"], key=lambda x: x.seq
        )
        reflections_final = sorted(report.reflections, key=lambda x: x.seq)
        questions_final = sorted(report.questions, key=lambda x: x.seq)
        items.append(
            PlaybackQueueItem(
                id=report.id,
                title=report.title,
                speaker=report.speaker,
                meeting_time=report.meeting_time,
                script_final=report.script_final,
                highlights_final=[h.highlight_text for h in highlights_final][:2],
                reflections_final=[r.reflection_text for r in reflections_final][:5],
                questions_final=[q.question_text for q in questions_final][:3],
                question_persona=_normalize_question_persona_key(
                    report.question_persona
                ),
                localized=_build_localized_payload(
                    db,
                    report,
                    [h.highlight_text for h in highlights_final][:2],
                    [r.reflection_text for r in reflections_final][:5],
                    [q.question_text for q in questions_final][:3],
                    include_audio=include_audio,
                    audio_languages=audio_languages,
                ),
            )
        )

    return PlaybackQueueResponse(items=items, total=len(items))
